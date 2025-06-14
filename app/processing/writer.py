import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PySide6 import QtWidgets
from datetime import datetime, date, time

def save_result(result: dict, log: list):
    path, _ = QtWidgets.QFileDialog.getSaveFileName(
        None, "Сохранить файл", "", "Excel (*.xlsx)"
    )
    if not path:
        return

    if os.path.exists(path):
        r = QtWidgets.QMessageBox.question(
            None, "Перезапись", "Файл существует. Перезаписать?"
        )
        if r != QtWidgets.QMessageBox.Yes:
            return

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for sheet_name, df in result.items():
            if df.shape[1] == 0:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col[:4])
            letter  = get_column_letter(col[0].column)
            ws.column_dimensions[letter].width = max_len + 2

            for cell in col[1:]: 
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = 'DD.MM.YYYY'
                elif isinstance(cell.value, time):
                    cell.number_format = 'HH:MM:SS'
    wb.save(path)

    log_path = os.path.splitext(path)[0] + ".log"
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(log))

    QtWidgets.QMessageBox.information(
        None, "Готово", f"Сохранено: {path}\nЛог: {log_path}"
    )
