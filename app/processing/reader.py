import os
import re
import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz
from PySide6 import QtWidgets
from typing import Tuple, Dict, List
from app.config import load_rules
from app.processing.transformer import (
    apply_word_replace,
    apply_word_filter,
    extract_units_to_headers,
    apply_column_rules,
    apply_unit_conversions,
)

def process_files(
    paths: List[str],
    rules: dict
) -> Tuple[Dict[str, pd.DataFrame], List[str]]:
    log: List[str] = []
    all_sheets: Dict[str, List[pd.DataFrame]] = {}
    moved_sheets: Dict[str, List[pd.DataFrame]] = {}

    for p in paths:
        engine = 'xlrd' if p.lower().endswith('.xls') else None
        log.append(f"Чтение файла {os.path.basename(p)}")
        sheets = pd.read_excel(p, sheet_name=None, header=None, engine=engine)
        wb = load_workbook(p, data_only=True)

        for sh, raw in sheets.items():
            ws = wb[sh]
            df = _detect_and_fix_header(raw, ws, rules, log)
            df = _remove_duplicate_header_rows(df)
            new_name = _map_sheet_name(sh, rules, log)
            log.append(f"Лист «{sh}» → «{new_name}»")

            if rules["column_word_filter"].get("enabled", True):
                core, tails = _split_rows_by_keywords(
                    df,
                    rules["column_word_filter"]["rules"],
                    log
                )
            else:
                core, tails = df, {}

            all_sheets.setdefault(new_name, []).append(core)
            for suffix, tail_df in tails.items():
                key = f"{new_name}_{suffix}"
                moved_sheets.setdefault(key, []).append(tail_df)

    sheet_cfg = rules["sheet_rules"]
    if sheet_cfg.get("enabled", True):
        clustered: Dict[str, List[pd.DataFrame]] = {}
        used = set()
        threshold = sheet_cfg.get("threshold", 90)
        auto_merge = sheet_cfg.get("auto_merge", True)

        for name, dfs in list(all_sheets.items()):
            if name in used:
                continue
            used.add(name)
            group = dfs[:]
            for other, other_dfs in list(all_sheets.items()):
                if other in used:
                    continue
                score = fuzz.token_set_ratio(name.lower(), other.lower())
                if score >= threshold:
                    if auto_merge:
                        log.append(f"FuzzyWuzzy: объединение '{other}' → '{name}' ({round(score)}%)")
                        group.extend(other_dfs)
                        used.add(other)
                    else:
                        ans = QtWidgets.QMessageBox.question(
                            None,
                            "Объединить листы?",
                            f"Объединить листы '{other}' → '{name}' ({score}%)?",
                            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                            QtWidgets.QMessageBox.Yes
                        )
                        if ans == QtWidgets.QMessageBox.Yes:
                            log.append(f"Пользователь подтвердил слияние '{other}' → '{name}'")
                            group.extend(other_dfs)
                            used.add(other)
            clustered[name] = group
        all_sheets = clustered

    result: Dict[str, pd.DataFrame] = {}
    for name, dfs in all_sheets.items():
        if len(dfs) > 1:
            log.append(f"Объединение {len(dfs)} частей листа «{name}»")
        dfs = [__ensure_unique_columns(df) for df in dfs]
        merged = pd.concat(dfs, ignore_index=True, sort=False)

        col_source = {}
        for idx, df in enumerate(dfs):
            for col in df.columns:
                col_source.setdefault(col, idx)

        if rules["word_replace"].get("enabled", True):
            merged = apply_word_replace(merged, rules["word_replace"], log)

        if rules["word_filter"].get("enabled", True):
            merged = apply_word_filter(merged, rules["word_filter"], log)

        merged = extract_units_to_headers(merged, rules, log)

        if rules["column_rules"].get("enabled", True):
            merged = apply_column_rules(merged, rules["column_rules"], log, col_source)

        if rules["unit_rules"].get("enabled", True):
            merged = apply_unit_conversions(merged, rules["unit_rules"], log)

        merged = merged.dropna(how="all").reset_index(drop=True)
        result[name] = merged

    for sheet_name, tails in moved_sheets.items():
        log.append(f"Перенесены строки в лист «{sheet_name}»")
        result[sheet_name] = pd.concat(tails, ignore_index=True, sort=False)

    return result, log

def __ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = df.columns.tolist()
    counts = {}
    new_cols = []
    for c in cols:
        if c in counts:
            counts[c] += 1
            new_cols.append(f"{c}.{counts[c]}")
        else:
            counts[c] = 0
            new_cols.append(c)
    df.columns = new_cols
    return df

def _split_rows_by_keywords(
    df: pd.DataFrame,
    rules: List[dict],
    log: List[str]
) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    main_df = df.copy().reset_index(drop=True)
    moved: Dict[str, pd.DataFrame] = {}

    for rule in rules:
        word = rule["word"].strip()
        delete_row = bool(rule.get("delete_row", False))
        mask = main_df.apply(
            lambda row: row.astype(str)
                .str.contains(rf"\b{re.escape(word)}\b", case=False, regex=True)
                .any(),
            axis=1
        )
        if not mask.any():
            continue

        first_idx = int(mask.idxmax())
        if delete_row:
            main_df = main_df.iloc[:first_idx].reset_index(drop=True)
        else:
            moved[word] = main_df.iloc[first_idx:].reset_index(drop=True)
            main_df = main_df.iloc[:first_idx].reset_index(drop=True)

    return main_df, moved

def _detect_and_fix_header(raw: pd.DataFrame, ws, rules: dict, log: List[str]) -> pd.DataFrame:
    def get_merged_value(ws, row: int, col: int):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            return cell.value
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return ws.cell(row=merged.min_row, column=merged.min_col).value
        return None

    header_start = 0
    while header_start < raw.shape[0] and raw.iloc[header_start].isna().all():
        header_start += 1
    def letter_ratio(row):
        total = len(row)
        letters = sum(isinstance(v, str) and any(ch.isalpha() for ch in v) for v in row)
        return letters / total

    header_rows = []
    has_any_merge = False
    for idx in range(header_start, min(header_start + 10, raw.shape[0])):
        ws_row = idx + 1
        row = raw.iloc[idx]
        has_hmerge = any(
            m.min_row == ws_row and m.min_col != m.max_col
            for m in ws.merged_cells.ranges
        )
        has_text = letter_ratio(row) > 0.5
        if has_hmerge or has_text:
            header_rows.append(idx)
            if has_hmerge:
                has_any_merge = True
        else:
            break

    if len(header_rows) > 1:
        HEADER_LEVEL_SEPARATOR = "; "
        if not has_any_merge:
            header_rows = []
        else:
            ncols = raw.shape[1]
            levels = []
            for idx in header_rows:
                ws_row = idx + 1
                vals = []
                for c in range(1, ncols + 1):
                    v = get_merged_value(ws, ws_row, c)
                    vals.append("" if v is None else str(v).strip())
                levels.append(vals)
            flat = []
            for ci in range(ncols):
                raw_parts = [levels[r][ci] for r in range(len(levels)) if levels[r][ci]]
                parts = []
                for p in raw_parts:
                    if not parts or p != parts[-1]:
                        parts.append(p)
                flat.append(HEADER_LEVEL_SEPARATOR.join(parts) if parts else f"col{ci}")

            data = raw.iloc[header_rows[-1] + 1 :].reset_index(drop=True)
            data.columns = flat
            start, end = header_start + 1, header_rows[-1] + 1
            log.append(f"Иерархический заголовок: строки {start}–{end}")
            return data

    if raw.shape[0] < 1 or raw.shape[1] < 1:
        return raw.copy()

    best_i = 0
    best_s = 0
    cols = raw.shape[1]
    skip_keywords = rules.get('skip_rows_keywords', [])

    for i in range(min(10, raw.shape[0])):
        row = raw.iloc[i]
        if any(
            isinstance(v, str) and any(kw.lower() in v.lower() for kw in skip_keywords)
            for v in row.tolist()
        ):
            continue

        cnt = sum(1 for v in row.tolist() if isinstance(v, str) and v.strip())
        if cnt / cols > best_s:
            best_s = cnt / cols
            best_i = i

    if best_i > 0 and best_s > 0.6:
        header = raw.iloc[best_i].tolist()
        df = raw.iloc[best_i + 1 :].reset_index(drop=True)
        df.columns = header
        log.append(f"Заголовок найден в строке {best_i}: {header}")
        return df

    header = raw.iloc[0].tolist()
    df = raw.iloc[1:].reset_index(drop=True)
    df.columns = header
    log.append(f"Использован первый ряд как заголовок: {header}")
    return df

def _remove_duplicate_header_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.shape[1] == 0:
        return df
    first_col_name = df.columns[0]
    anchor_str = str(first_col_name).strip().lower()
    mask = (
        df.iloc[:, 0]
          .astype(str)
          .str.strip()
          .str.lower()
          .eq(anchor_str)
    )
    return df.loc[~mask].reset_index(drop=True)

def _map_sheet_name(name: str, rules: dict, log: List[str]) -> str:
    cfg = rules["sheet_rules"]
    lname = name.strip().lower()

    for rule in cfg.get("rules", []):
        variants = {rule["target"].lower()} | {s.lower() for s in rule.get("synonyms", [])}
        if lname in variants:
            return name if rule.get("no_merge", False) else rule["target"]

    best_score, best_target = 0, None
    for rule in cfg.get("rules", []):
        for cand in (rule["target"], *rule.get("synonyms", [])):
            sc = fuzz.token_set_ratio(lname, cand.lower())
            if sc > best_score:
                best_score, best_target = sc, rule["target"]

    if best_score >= cfg.get("threshold", 90) and best_target:
        if cfg.get("auto_merge", True):
            log.append(f"FuzzyWuzzy: '{name}' → '{best_target}' ({round(best_score)}%)")
            return best_target
        else:
            return best_target

    return name
